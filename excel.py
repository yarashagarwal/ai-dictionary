from typing import List, Dict, Optional, Literal, Union
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import os

def excel_action(
    action: Literal['read', 'write', 'append'],
    data: Optional[List[Dict[str, str]]] = None,
    filename: str = 'dict.xlsx',
    sheet_name: str = 'norsk'
) -> Union[List[Dict[str, str]], bool]:
    """
    Perform an action on the Excel file dict.xlsx using openpyxl.
    action: 'read', 'write', or 'append'
    data: For 'write' or 'append', should be a list of dicts with keys:
        'word', 'meaning', 'example', 'example_english'
    filename: Excel file name (default: dict.xlsx)
    sheet_name: Sheet to operate on (default: norsk)
    Returns: For 'read', returns all rows as a list of dicts
    """
    headers = ['word', 'meaning', 'example', 'example_english']
    if action == 'read':
        if not os.path.exists(filename):
            return []
        wb = openpyxl.load_workbook(filename)
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        # Assume first row is header
        header = rows[0]
        return [dict(zip(header, row)) for row in rows[1:]]

    elif action == 'write':
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        ws.append(headers)
        if data:
            for entry in data:
                ws.append([entry.get(h, '') for h in headers])
        wb.save(filename)
        return True

    elif action == 'append':
        if os.path.exists(filename):
            wb = openpyxl.load_workbook(filename)
            ws = wb[sheet_name]
            # Check if header exists
            if ws.max_row == 0:
                ws.append(headers)
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name
            ws.append(headers)
        if data:
            for entry in data:
                ws.append([entry.get(h, '') for h in headers])
        wb.save(filename)
        return True

    else:
        raise ValueError("Invalid action. Use 'read', 'write', or 'append'.")